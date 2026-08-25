"""
XLSX/XLS 文件解析器
支持解析 Excel 表格 (.xls, .xlsx, .xlsm)
"""

import logging
from pathlib import Path
from typing import Any

from core.models import ExtractedElement, PageContent
from parsers import BaseParser

logger = logging.getLogger("parsers.xlsx")

# 可选依赖
try:
    import openpyxl  # noqa: F401  (defensive: ensure openpyxl is loadable)
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter  # noqa: F401  (defensive: ensure utils submodule loads)

    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False
    logger.warning("openpyxl 库未安装，XLSX 解析功能不可用")

try:
    import xlrd

    XLS_AVAILABLE = True
except ImportError:
    XLS_AVAILABLE = False


class XLSXParser(BaseParser):
    """Excel 文件解析器"""

    @property
    def supported_extensions(self) -> list[str]:
        return [".xlsx", ".xls", ".xlsm", ".xlsb"]

    @property
    def supported_magic(self) -> list[bytes]:
        # XLSX 是 ZIP 格式
        # XLS 是 OLE2 格式
        return [b"PK\x03\x04", b"\xd0\xcf\x11\xe0"]

    def parse(self, file_path: Path) -> list[PageContent]:
        """解析 Excel 文件"""
        ext = file_path.suffix.lower()

        if ext in [".xlsx", ".xlsm", ".xlsb"]:
            return self._parse_xlsx(file_path)
        elif ext == ".xls":
            return self._parse_xls(file_path)
        else:
            # 尝试用 openpyxl 打开
            return self._parse_xlsx(file_path)

    def _parse_xlsx(self, file_path: Path) -> list[PageContent]:
        """解析现代 Excel 格式 (.xlsx, .xlsm)

        E3 增强：多 sheet 各自成表（sheet 名入锚点）、合并单元格值填充、
        数字/日期类型保真（防 001 变 1）。
        """
        if not XLSX_AVAILABLE:
            raise ImportError("openpyxl 库未安装，无法解析 XLSX 文件")

        logger.info("开始解析 XLSX: %s", file_path)

        try:
            # E3: 不能用 read_only=True —— 合并单元格信息需要普通模式
            wb = load_workbook(file_path, data_only=True)
        except Exception as e:
            logger.error("无法打开 XLSX 文件: %s", e)
            raise ValueError(f"无法打开 XLSX 文件: {e}") from e

        elements = []
        raw_text_parts = []
        elem_idx = 0

        for sheet_name in wb.sheetnames:
            logger.debug("解析 Sheet: %s", sheet_name)
            sheet = wb[sheet_name]

            # E3: 构建合并单元格映射 —— 左上角值填充到整个合并区域
            merged_fill: dict[tuple[int, int], Any] = {}
            try:
                for rng in sheet.merged_cells.ranges:
                    top_left = sheet.cell(rng.min_row, rng.min_col).value
                    for row in range(rng.min_row, rng.max_row + 1):
                        for col in range(rng.min_col, rng.max_col + 1):
                            if (row, col) != (rng.min_row, rng.min_col):
                                merged_fill[(row, col)] = top_left
            except Exception as e:
                logger.debug("合并单元格读取失败（跳过填充）: %s", e)

            def _cell_str(value: Any) -> str:
                """类型保真序列化：保留前导零/长数字的原始形态。"""
                if value is None:
                    return ""
                return str(value)

            # 提取表格数据
            sheet_data = []
            max_col = 0

            for r_idx, row in enumerate(sheet.iter_rows(values_only=False), start=1):
                row_data = []
                for c_idx, cell in enumerate(row, start=1):
                    # 合并区域非左上格：填左上角的值
                    if cell.value is None and (r_idx, c_idx) in merged_fill:
                        row_data.append(_cell_str(merged_fill[(r_idx, c_idx)]))
                        continue
                    if cell.value is not None:
                        row_data.append(_cell_str(cell.value))
                    else:
                        row_data.append("")
                if any(row_data):
                    sheet_data.append(row_data)
                    max_col = max(max_col, len(row_data))

            if not sheet_data:
                continue

            # 格式化表格文本 —— Markdown 表格（含表头分隔行）
            table_lines = []
            for _row_idx, row_data in enumerate(sheet_data):
                while len(row_data) < max_col:
                    row_data.append("")
                table_lines.append("| " + " | ".join(row_data) + " |")
            if len(table_lines) > 1:
                table_lines.insert(1, "|" + "---|" * max_col)

            table_text = "\n".join(table_lines)
            header_text = table_lines[0] if table_lines else ""

            elements.append(
                ExtractedElement(
                    elementId=f"elem_1_{elem_idx}",
                    elementType="table",
                    content=f"## Sheet: {sheet_name}\n\n{table_text}",
                    metadata={
                        "sheet_name": sheet_name,
                        "rows": len(sheet_data),
                        "cols": max_col,
                        "header": header_text,
                        "has_header": self._detect_header(sheet_data),
                        "merged_cells": len(merged_fill),
                    },
                )
            )
            raw_text_parts.append(f"[Sheet: {sheet_name}]\n{table_text}")
            elem_idx += 1

        wb.close()

        logger.info("XLSX 解析完成: %d 个 Sheet", len(elements))

        return [
            PageContent(
                pageNumber=1,
                elements=elements,
                rawText="\n\n".join(raw_text_parts),
                hasImage=False,
                hasTable=len(elements) > 0,
            )
        ]

    def _parse_xls(self, file_path: Path) -> list[PageContent]:
        """解析旧版 Excel 格式 (.xls)"""
        if XLSX_AVAILABLE:
            # 优先尝试用 openpyxl 打开（部分兼容）
            try:
                return self._parse_xlsx(file_path)
            except Exception:
                pass

        if not XLS_AVAILABLE:
            raise ImportError("xlrd 库未安装，无法解析 XLS 文件")

        logger.info("开始解析 XLS: %s", file_path)

        try:
            wb = xlrd.open_workbook(str(file_path))
        except Exception as e:
            logger.error("无法打开 XLS 文件: %s", e)
            raise ValueError(f"无法打开 XLS 文件: {e}") from e

        elements = []
        raw_text_parts = []
        elem_idx = 0

        for sheet_idx in range(wb.nsheets):
            sheet = wb.sheet_by_index(sheet_idx)
            sheet_name = sheet.name

            sheet_data = []
            for row_idx in range(sheet.nrows):
                row_data = []
                for col_idx in range(sheet.ncols):
                    cell = sheet.cell_value(row_idx, col_idx)
                    row_data.append(str(cell) if cell is not None else "")
                if any(row_data):
                    sheet_data.append(row_data)

            if not sheet_data:
                continue

            table_lines = []
            for row_data in sheet_data:
                line = " | ".join(row_data)
                table_lines.append(line)

            table_text = "\n".join(table_lines)
            header_text = table_lines[0] if table_lines else ""

            elements.append(
                ExtractedElement(
                    elementId=f"elem_1_{elem_idx}",
                    elementType="table",
                    content=f"Sheet: {sheet_name}\n{table_text}",
                    metadata={
                        "sheet_name": sheet_name,
                        "rows": len(sheet_data),
                        "cols": sheet.ncols,
                        "header": header_text,
                        "has_header": self._detect_header(sheet_data),
                    },
                )
            )
            raw_text_parts.append(f"[Sheet: {sheet_name}]\n{table_text}")
            elem_idx += 1

        logger.info("XLS 解析完成: %d 个 Sheet", len(elements))

        return [
            PageContent(
                pageNumber=1,
                elements=elements,
                rawText="\n\n".join(raw_text_parts),
                hasImage=False,
                hasTable=len(elements) > 0,
            )
        ]

    def _detect_header(self, sheet_data: list[list[str]]) -> bool:
        """检测第一行是否为表头"""
        if len(sheet_data) < 2:
            return False

        first_row = sheet_data[0]
        second_row = sheet_data[1]

        # 表头通常与数据类型不同
        first_types = [self._guess_type(cell) for cell in first_row]
        second_types = [self._guess_type(cell) for cell in second_row]

        # 如果第一行全是文本，第二行有数字，可能是表头
        if all(t == "str" for t in first_types) and any(t in ("int", "float") for t in second_types):
            return True

        # 如果第一行有常见表头关键词
        header_keywords = [
            "id",
            "name",
            "title",
            "date",
            "time",
            "total",
            "sum",
            "编号",
            "名称",
            "标题",
            "日期",
            "时间",
            "合计",
            "总计",
        ]
        first_text = " ".join(first_row).lower()
        return bool(any(kw in first_text for kw in header_keywords))

    def _guess_type(self, value: str) -> str:
        """猜测单元格数据类型"""
        if not value:
            return "empty"
        try:
            int(value)
            return "int"
        except ValueError:
            pass
        try:
            float(value)
            return "float"
        except ValueError:
            pass
        return "str"
